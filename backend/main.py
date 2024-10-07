from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
import pandas as pd
import openpyxl
from io import BytesIO

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173"],  # Adjust this to match your Vue.js dev server
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

def get_cell_value(cell):
    if cell.data_type == 'f':
        return cell.value
    return cell.value

@app.post("/upload-excel/")
async def upload_excel(file: UploadFile = File(...)):
    contents = await file.read()
    try:
        if not file.filename.endswith('.xlsx'):
            raise HTTPException(status_code=400, detail="Unsupported file format. Please upload an .xlsx file.")

        # Use openpyxl to read the file and extract formulas
        workbook = openpyxl.load_workbook(BytesIO(contents), data_only=False)
        sheets_data = {}
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            data = []
            for row in sheet.iter_rows(values_only=False):
                data.append([get_cell_value(cell) for cell in row])
            
            # Convert to pandas DataFrame for efficient handling
            df = pd.DataFrame(data[1:], columns=data[0])
            
            sheets_data[sheet_name] = {
                "column_names": df.columns.tolist(),
                "data": df.to_dict(orient='records')
            }
        
        response = {
            "sheets": sheets_data,
            "sheet_names": workbook.sheetnames
        }
        return response
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not process file: {str(e)}")

@app.get("/")
async def root():
    return {"message": "Hello World"}